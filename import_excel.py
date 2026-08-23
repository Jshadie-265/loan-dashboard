"""Import and consolidate one or more Loan.xlsx workbooks into SQLite.

Imports are safe to repeat. Identical records with the same source ID are
skipped, while conflicting IDs from another workbook receive a new local ID so
that neither file's data is lost. Repayments are remapped to their imported
loan when a loan ID conflict occurs.
"""

import sys
from datetime import date, datetime
from pathlib import Path
from typing import BinaryIO, Iterable

import openpyxl

from business_logic import calculate_due_date, compute_loan_terms
from db import get_connection, init_db, next_id


def to_date(value) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value:
        try:
            return datetime.fromisoformat(str(value)).date().isoformat()
        except ValueError:
            return None
    return None


def data_rows(ws, id_column: int, fallback_column: int):
    """Yield filled register rows; formulas without cached IDs use a meaningful field."""
    for row in range(4, ws.max_row + 1):
        if ws.cell(row, id_column).value is not None or ws.cell(row, fallback_column).value is not None:
            yield row


def source_id(value_sheet, formula_sheet, row: int, prefix: str) -> str:
    """Return a saved ID or recreate the workbook's generated row-based ID."""
    value = text(value_sheet.cell(row, 1).value)
    if value:
        return value
    formula = formula_sheet.cell(row, 1).value if formula_sheet else None
    if isinstance(formula, str) and formula.startswith("="):
        return f"{prefix}{row - 3:03d}"
    return ""


def text(value) -> str:
    return str(value or "").strip()


def same_text(left, right) -> bool:
    return text(left).casefold() == text(right).casefold()


def find_customer_id(conn, full_name: str, national_id: str | None = None) -> str | None:
    """Find a borrower by national ID when present, otherwise by their name."""
    if text(national_id):
        row = conn.execute("SELECT customer_id FROM customers WHERE lower(COALESCE(national_id, '')) = lower(?)", (text(national_id),)).fetchone()
        if row:
            return row["customer_id"]
    row = conn.execute("SELECT customer_id FROM customers WHERE lower(full_name) = lower(?)", (text(full_name),)).fetchone()
    if row:
        return row["customer_id"]
    return None


def get_or_create_customer(conn, full_name: str, national_id: str | None = None) -> str:
    customer_id = find_customer_id(conn, full_name, national_id)
    if customer_id:
        return customer_id
    customer_id = next_id(conn, "customers", "customer_id", "CUST-")
    conn.execute("INSERT INTO customers (customer_id, full_name, national_id) VALUES (?, ?, ?)", (customer_id, text(full_name), text(national_id) or None))
    return customer_id


def import_customers(conn, ws, formula_ws=None) -> int:
    inserted = 0
    for row in data_rows(ws, 1, 2):
        full_name = text(ws.cell(row, 2).value)
        national_id = text(ws.cell(row, 4).value)
        if not full_name:
            continue
        requested_id = source_id(ws, formula_ws, row, "C") or next_id(conn, "customers", "customer_id", "CUST-")
        existing_by_id = conn.execute("SELECT full_name, national_id FROM customers WHERE customer_id=?", (requested_id,)).fetchone()
        if existing_by_id and (same_text(existing_by_id["full_name"], full_name) or (national_id and same_text(existing_by_id["national_id"], national_id))):
            continue
        if find_customer_id(conn, full_name, national_id):
            # This borrower already exists under another source ID.
            continue
        customer_id = requested_id if existing_by_id is None else next_id(conn, "customers", "customer_id", "CUST-")
        status = text(ws.cell(row, 8).value or "Active").title()
        status = status if status in {"Active", "Inactive"} else "Active"
        conn.execute(
            """INSERT INTO customers (customer_id, full_name, phone, national_id, address, occupation, collateral, status, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (customer_id, full_name, ws.cell(row, 3).value, national_id or None, ws.cell(row, 5).value,
             ws.cell(row, 6).value, ws.cell(row, 7).value, status, ws.cell(row, 9).value),
        )
        inserted += 1
    return inserted


def capital_record_matches(row, txn_date, txn_type, description, money_in, money_out) -> bool:
    return (
        row is not None
        and row["date"] == txn_date
        and same_text(row["type"], txn_type)
        and same_text(row["description"], description)
        and float(row["money_in"]) == money_in
        and float(row["money_out"]) == money_out
    )


def find_matching_capital(conn, txn_date, txn_type, description, money_in, money_out):
    for row in conn.execute("SELECT * FROM capital"):
        if capital_record_matches(row, txn_date, txn_type, description, money_in, money_out):
            return row["transaction_id"]
    return None


def import_capital(conn, ws, formula_ws=None) -> int:
    inserted = 0
    for row in data_rows(ws, 1, 2):
        txn_date = to_date(ws.cell(row, 2).value)
        money_in, money_out = float(ws.cell(row, 5).value or 0), float(ws.cell(row, 6).value or 0)
        if not txn_date or (money_in <= 0 and money_out <= 0):
            continue
        if money_in > 0 and money_out > 0:
            money_out = 0.0
        txn_type, description = text(ws.cell(row, 3).value or "Other"), text(ws.cell(row, 4).value)
        requested_id = source_id(ws, formula_ws, row, "T") or next_id(conn, "capital", "transaction_id", "T")
        existing = conn.execute("SELECT * FROM capital WHERE transaction_id=?", (requested_id,)).fetchone()
        if capital_record_matches(existing, txn_date, txn_type, description, money_in, money_out):
            continue
        if existing and find_matching_capital(conn, txn_date, txn_type, description, money_in, money_out):
            continue
        transaction_id = requested_id if existing is None else next_id(conn, "capital", "transaction_id", "T")
        conn.execute("INSERT INTO capital (transaction_id, date, type, description, money_in, money_out) VALUES (?, ?, ?, ?, ?, ?)",
                     (transaction_id, txn_date, txn_type, description or None, money_in, money_out))
        inserted += 1
    return inserted


def loan_record_matches(row, customer_id, issued, period, due_date, principal, rate, interest, total_due, bank, account, reference, notes) -> bool:
    return (
        row is not None and row["customer_id"] == customer_id and row["date_taken"] == issued
        and same_text(row["period"], period) and row["due_date"] == due_date
        and float(row["principal"]) == principal and float(row["rate"]) == rate
        and float(row["interest"]) == interest and float(row["total_due"]) == total_due
        and same_text(row["bank"], bank) and same_text(row["account"], account)
        and same_text(row["reference"], reference) and same_text(row["notes"], notes)
    )


def find_matching_loan(conn, *values):
    for row in conn.execute("SELECT * FROM loans"):
        if loan_record_matches(row, *values):
            return row["loan_id"]
    return None


def import_loans(conn, ws, formula_ws=None) -> tuple[int, dict[str, str]]:
    """Import loans and return source-to-local IDs for repayment reconciliation."""
    inserted, loan_ids = 0, {}
    for row in data_rows(ws, 1, 2):
        borrower, issued = text(ws.cell(row, 2).value), to_date(ws.cell(row, 3).value)
        principal = float(ws.cell(row, 6).value or 0)
        if not borrower or not issued or principal <= 0:
            continue
        source_loan_id = source_id(ws, formula_ws, row, "L")
        period = text(ws.cell(row, 4).value or "Custom")
        due_date = to_date(ws.cell(row, 5).value)
        if not due_date and period != "Custom":
            try:
                due_date = calculate_due_date(date.fromisoformat(issued), period).isoformat()
            except ValueError:
                due_date = issued
        due_date = due_date or issued
        rate = float(ws.cell(row, 7).value or 0)
        calculated_interest, calculated_total = compute_loan_terms(principal, rate)
        interest = float(ws.cell(row, 8).value or calculated_interest)
        total_due = float(ws.cell(row, 9).value or calculated_total)
        bank, account, reference, notes = ws.cell(row, 13).value, ws.cell(row, 14).value, ws.cell(row, 15).value, ws.cell(row, 16).value
        customer_id = get_or_create_customer(conn, borrower)
        requested_id = source_loan_id or next_id(conn, "loans", "loan_id", "L")
        existing = conn.execute("SELECT * FROM loans WHERE loan_id=?", (requested_id,)).fetchone()
        values = (customer_id, issued, period, due_date, principal, rate, interest, total_due, bank, account, reference, notes)
        if loan_record_matches(existing, *values):
            loan_ids[source_loan_id or requested_id] = requested_id
            continue
        if existing and (matching_loan_id := find_matching_loan(conn, *values)):
            loan_ids[source_loan_id or requested_id] = matching_loan_id
            continue
        loan_id = requested_id if existing is None else next_id(conn, "loans", "loan_id", "L")
        conn.execute(
            """INSERT INTO loans (loan_id, customer_id, date_taken, period, due_date, principal, rate, interest, total_due, bank, account, reference, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (loan_id, customer_id, issued, period, due_date, principal, rate, interest, total_due, bank,
             str(account) if account is not None else None, reference, notes),
        )
        loan_ids[source_loan_id or requested_id] = loan_id
        inserted += 1
    return inserted, loan_ids


def repayment_record_matches(row, loan_id, payment_date, amount_paid, payment_method, reference, notes, recorded_by) -> bool:
    return (
        row is not None and row["loan_id"] == loan_id and row["payment_date"] == payment_date
        and float(row["amount_paid"]) == amount_paid and same_text(row["payment_method"], payment_method)
        and same_text(row["reference"], reference) and same_text(row["notes"], notes)
        and same_text(row["recorded_by"], recorded_by)
    )


def find_matching_repayment(conn, *values):
    for row in conn.execute("SELECT * FROM repayments"):
        if repayment_record_matches(row, *values):
            return row["payment_id"]
    return None


def import_repayments(conn, ws, loan_ids: dict[str, str], formula_ws=None) -> int:
    inserted = 0
    for row in data_rows(ws, 1, 2):
        source_loan_id, paid_on = text(ws.cell(row, 2).value), to_date(ws.cell(row, 4).value)
        amount = float(ws.cell(row, 5).value or 0)
        if not source_loan_id or not paid_on or amount <= 0:
            continue
        loan_id = loan_ids.get(source_loan_id, source_loan_id)
        if not conn.execute("SELECT 1 FROM loans WHERE loan_id=?", (loan_id,)).fetchone():
            continue
        source_payment_id = source_id(ws, formula_ws, row, "P")
        requested_id = source_payment_id or next_id(conn, "repayments", "payment_id", "PMT-")
        values = (loan_id, paid_on, amount, ws.cell(row, 6).value, ws.cell(row, 7).value, ws.cell(row, 8).value, ws.cell(row, 9).value)
        existing = conn.execute("SELECT * FROM repayments WHERE payment_id=?", (requested_id,)).fetchone()
        if repayment_record_matches(existing, *values):
            continue
        if existing and find_matching_repayment(conn, *values):
            continue
        payment_id = requested_id if existing is None else next_id(conn, "repayments", "payment_id", "PMT-")
        conn.execute(
            """INSERT INTO repayments (payment_id, loan_id, payment_date, amount_paid, payment_method, reference, notes, recorded_by)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (payment_id, *values),
        )
        inserted += 1
    return inserted


def import_workbook(source: str | Path | BinaryIO) -> dict[str, int]:
    """Import a compatible workbook and consolidate it with the current database."""
    init_db()
    if hasattr(source, "seek"):
        source.seek(0)
    workbook = openpyxl.load_workbook(source, data_only=True)
    if hasattr(source, "seek"):
        source.seek(0)
    formula_workbook = openpyxl.load_workbook(source, data_only=False)
    required = {"Capital", "Loans", "Repayments"}
    missing = required.difference(workbook.sheetnames)
    if missing:
        raise ValueError(f"Workbook is missing required sheet(s): {', '.join(sorted(missing))}")
    conn = get_connection()
    try:
        with conn:
            customers = import_customers(conn, workbook["Customers"], formula_workbook["Customers"]) if "Customers" in workbook.sheetnames else 0
            capital = import_capital(conn, workbook["Capital"], formula_workbook["Capital"])
            loans, loan_ids = import_loans(conn, workbook["Loans"], formula_workbook["Loans"])
            repayments = import_repayments(conn, workbook["Repayments"], loan_ids, formula_workbook["Repayments"])
        return {"customers": customers, "capital": capital, "loans": loans, "repayments": repayments}
    finally:
        conn.close()


def import_workbooks(sources: Iterable[str | Path | BinaryIO]) -> dict[str, int]:
    """Import several workbooks in sequence and return one consolidated summary."""
    totals = {"customers": 0, "capital": 0, "loans": 0, "repayments": 0}
    for source in sources:
        results = import_workbook(source)
        for name, count in results.items():
            totals[name] += count
    return totals


def main(xlsx_path: str) -> None:
    results = import_workbook(xlsx_path)
    print("Import complete: " + ", ".join(f"{count} {name}" for name, count in results.items()))


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python import_excel.py path/to/Loan.xlsx")
        raise SystemExit(1)
    main(sys.argv[1])
